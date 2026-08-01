"""Tests for the load-spec-driven workbook generator.

The centerpiece is the round-trip: build a synthetic EIB template (multi-row
header band, repeating-group sheet), generate a filled load from CSVs via a
YAML spec, then parse and validate the output — it must come back clean,
with types and spreadsheet keys the validator finds nothing wrong with.
All names and IDs are synthetic.
"""

import datetime as dt
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook as XlsxWorkbook
from openpyxl import load_workbook

from eib_toolkit.generate import (
    ColumnMapping,
    GenerateError,
    LoadSpec,
    SheetMapping,
    generate_workbook,
    load_spec,
)
from eib_toolkit.parser import parse_workbook
from eib_toolkit.validate import validate_workbook


def _write_xlsx(path: Path, sheets: dict[str, list[list[Any]]]) -> Path:
    wb = XlsxWorkbook()
    assert wb.active is not None
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    wb.save(path)
    return path


def _template(path: Path) -> Path:
    """Synthetic two-data-sheet EIB template: workers + repeating allowances."""
    return _write_xlsx(
        path,
        {
            "Instructions": [
                ["Synthetic EIB template for tests. All names and IDs are fake."],
            ],
            "Workers": [
                [None, None, "Personal Data", None, "Compensation", None],
                [
                    "Spreadsheet Key*",
                    "Employee_Reference",
                    "Legal Name",
                    "Hire Date",
                    "Annual Amount",
                    "Active",
                ],
                [
                    "Numeric. Required.",
                    "Reference ID Type: Employee_ID",
                    "Text. Required.",
                    "Date (yyyy-mm-dd)",
                    "Numeric. Optional.",
                    "Boolean (Y/N)",
                ],
            ],
            "Allowances": [
                ["Spreadsheet Key*", "Allowance Plan", "Amount"],
                ["Numeric. Required.", "Text. Required.", "Numeric. Required."],
            ],
        },
    )


def _write_csvs(tmp_path: Path) -> None:
    (tmp_path / "workers.csv").write_text(
        "employee_id,name,hire_date,amount,active\n"
        "EMP-001,Avery Example,2026-01-15,55000,yes\n"
        "EMP-002,Blake Sample,2026-02-01,61000.50,no\n",
        encoding="utf-8",
    )
    (tmp_path / "allowances.csv").write_text(
        "employee_id,plan,amount\n"
        "EMP-001,Transport,150\n"
        "EMP-001,Meal,80\n"
        "EMP-002,Transport,150\n",
        encoding="utf-8",
    )


_SPEC_YAML = """\
sheets:
  - sheet: Workers
    source: workers.csv
    key_from: employee_id
    columns:
      - column: Employee_Reference
        source: employee_id
      - column: Legal Name
        source: name
      - column: Hire Date
        source: hire_date
      - column: Annual Amount
        source: amount
      - column: Active
        source: active
  - sheet: Allowances
    source: allowances.csv
    key_from: employee_id
    columns:
      - column: Allowance Plan
        source: plan
      - column: Amount
        source: amount
"""


def _generate(tmp_path: Path) -> tuple[Path, Any]:
    template = _template(tmp_path / "template.xlsx")
    _write_csvs(tmp_path)
    spec_path = tmp_path / "load.yaml"
    spec_path.write_text(_SPEC_YAML, encoding="utf-8")
    out = tmp_path / "out" / "filled.xlsx"
    report = generate_workbook(template, load_spec(spec_path), out)
    return out, report


class TestRoundTrip:
    def test_output_parses_and_validates_clean(self, tmp_path: Path) -> None:
        out, report = _generate(tmp_path)
        assert report.warnings == []
        assert report.rows_written == {"Workers": 2, "Allowances": 3}

        parsed = parse_workbook(out)
        validation = validate_workbook(parsed)
        assert validation.findings == []
        assert validation.ok

    def test_keys_assigned_sequentially_and_joined(self, tmp_path: Path) -> None:
        out, report = _generate(tmp_path)
        assert report.key_assignments == {"emp-001": 1, "emp-002": 2}

        parsed = parse_workbook(out)
        workers = parsed.sheet_data("Workers")
        allowances = parsed.sheet_data("Allowances")
        assert workers is not None and allowances is not None
        assert workers.column_values(0) == [1, 2]
        # EMP-001's two allowance rows share key 1; EMP-002's row has key 2.
        assert allowances.column_values(0) == [1, 1, 2]

    def test_values_written_with_real_types(self, tmp_path: Path) -> None:
        out, _ = _generate(tmp_path)
        wb = load_workbook(out)
        ws = wb["Workers"]
        # Data starts at row 4 (group row, field row, hint row above it).
        assert ws.cell(row=4, column=4).value == dt.datetime(2026, 1, 15)  # noqa: DTZ001
        assert ws.cell(row=4, column=5).value == 55000
        assert isinstance(ws.cell(row=4, column=5).value, int)
        assert ws.cell(row=5, column=5).value == pytest.approx(61000.50)
        assert ws.cell(row=4, column=6).value == "Y"  # "yes" normalized
        assert ws.cell(row=5, column=6).value == "N"
        wb.close()

    def test_header_band_preserved_from_template(self, tmp_path: Path) -> None:
        out, _ = _generate(tmp_path)
        parsed = parse_workbook(out)
        workers = parsed.template.sheet("Workers")
        assert workers is not None
        assert [c.header for c in workers.columns] == [
            "Spreadsheet Key",
            "Employee_Reference",
            "Legal Name",
            "Hire Date",
            "Annual Amount",
            "Active",
        ]
        assert workers.columns[1].ref_id_type == "Employee_ID"
        allowances = parsed.template.sheet("Allowances")
        assert allowances is not None
        assert allowances.role.value == "repeating"


class TestKeyStrategies:
    def test_from_source_writes_business_key_verbatim(self, tmp_path: Path) -> None:
        template = _template(tmp_path / "template.xlsx")
        _write_csvs(tmp_path)
        spec_path = tmp_path / "load.yaml"
        spec_path.write_text(_SPEC_YAML + "key_strategy: from_source\n", encoding="utf-8")
        out = tmp_path / "filled.xlsx"
        generate_workbook(template, load_spec(spec_path), out)

        parsed = parse_workbook(out)
        workers = parsed.sheet_data("Workers")
        assert workers is not None
        assert workers.column_values(0) == ["EMP-001", "EMP-002"]

    def test_duplicate_primary_business_key_warns_and_shares_key(
        self, tmp_path: Path
    ) -> None:
        template = _template(tmp_path / "template.xlsx")
        _write_csvs(tmp_path)
        (tmp_path / "workers.csv").write_text(
            "employee_id,name,hire_date,amount,active\n"
            "EMP-001,Avery Example,2026-01-15,55000,yes\n"
            "EMP-001,Avery Duplicate,2026-02-01,61000,no\n",
            encoding="utf-8",
        )
        (tmp_path / "allowances.csv").write_text(
            "employee_id,plan,amount\nEMP-001,Transport,150\n", encoding="utf-8"
        )
        spec_path = tmp_path / "load.yaml"
        spec_path.write_text(_SPEC_YAML, encoding="utf-8")
        report = generate_workbook(template, load_spec(spec_path), tmp_path / "f.xlsx")
        assert any("duplicate business key 'EMP-001'" in w for w in report.warnings)

    def test_orphan_business_key_leaves_key_empty_and_warns(self, tmp_path: Path) -> None:
        template = _template(tmp_path / "template.xlsx")
        _write_csvs(tmp_path)
        (tmp_path / "allowances.csv").write_text(
            "employee_id,plan,amount\nEMP-999,Transport,150\n", encoding="utf-8"
        )
        spec_path = tmp_path / "load.yaml"
        spec_path.write_text(_SPEC_YAML, encoding="utf-8")
        out = tmp_path / "filled.xlsx"
        report = generate_workbook(template, load_spec(spec_path), out)
        assert any("EMP-999" in w and "no matching primary row" in w for w in report.warnings)
        # The validator then reports the missing join key on that row.
        validation = validate_workbook(parse_workbook(out))
        assert "EIB051" in validation.by_code()


class TestCoercion:
    def test_bad_values_written_verbatim_and_warned(self, tmp_path: Path) -> None:
        template = _template(tmp_path / "template.xlsx")
        _write_csvs(tmp_path)
        (tmp_path / "workers.csv").write_text(
            "employee_id,name,hire_date,amount,active\n"
            "EMP-001,Avery Example,01/15/2026,fifty,maybe\n",
            encoding="utf-8",
        )
        (tmp_path / "allowances.csv").write_text(
            "employee_id,plan,amount\n", encoding="utf-8"
        )
        spec_path = tmp_path / "load.yaml"
        spec_path.write_text(_SPEC_YAML, encoding="utf-8")
        out = tmp_path / "filled.xlsx"
        report = generate_workbook(template, load_spec(spec_path), out)

        assert len(report.warnings) == 3  # date, numeric, boolean
        # Nothing silently dropped: the validator sees exactly those values.
        validation = validate_workbook(parse_workbook(out))
        codes = validation.by_code()
        assert "EIB011" in codes  # locale date
        assert "EIB020" in codes  # non-numeric
        assert "EIB030" in codes  # bad boolean

    def test_constants_applied_to_every_row(self, tmp_path: Path) -> None:
        template = _write_xlsx(
            tmp_path / "t.xlsx",
            {
                "Data": [
                    ["Employee_Reference", "Currency"],
                    ["Reference ID Type: Employee_ID", "Text."],
                ]
            },
        )
        (tmp_path / "rows.csv").write_text(
            "employee_id\nEMP-001\nEMP-002\n", encoding="utf-8"
        )
        spec = LoadSpec(
            sheets=[
                SheetMapping(
                    sheet="Data",
                    source="rows.csv",
                    columns=[
                        ColumnMapping(column="Employee_Reference", source="employee_id"),
                        ColumnMapping(column="Currency", const="USD"),
                    ],
                )
            ],
            base_dir=str(tmp_path),
        )
        out = tmp_path / "filled.xlsx"
        generate_workbook(template, spec, out)
        parsed = parse_workbook(out)
        data = parsed.sheet_data("Data")
        assert data is not None
        assert data.column_values(1) == ["USD", "USD"]


class TestSpecErrors:
    def _base(self, tmp_path: Path) -> tuple[Path, Path]:
        template = _template(tmp_path / "template.xlsx")
        _write_csvs(tmp_path)
        return template, tmp_path

    def _spec(self, tmp_path: Path, text: str) -> LoadSpec:
        path = tmp_path / "load.yaml"
        path.write_text(text, encoding="utf-8")
        return load_spec(path)

    def test_unknown_sheet(self, tmp_path: Path) -> None:
        template, base = self._base(tmp_path)
        spec = self._spec(
            base,
            "sheets:\n  - sheet: Nope\n    source: workers.csv\n    columns:\n"
            "      - column: Legal Name\n        source: name\n",
        )
        with pytest.raises(GenerateError, match="no data sheet named 'Nope'"):
            generate_workbook(template, spec, base / "f.xlsx")

    def test_unknown_column(self, tmp_path: Path) -> None:
        template, base = self._base(tmp_path)
        spec = self._spec(
            base,
            "sheets:\n  - sheet: Workers\n    source: workers.csv\n"
            "    key_from: employee_id\n    columns:\n"
            "      - column: Legal Name\n        source: name\n"
            "      - column: Shoe Size\n        source: name\n",
        )
        with pytest.raises(GenerateError, match="no column 'Shoe Size'"):
            generate_workbook(template, spec, base / "f.xlsx")

    def test_unmapped_required_column(self, tmp_path: Path) -> None:
        template, base = self._base(tmp_path)
        spec = self._spec(
            base,
            "sheets:\n  - sheet: Workers\n    source: workers.csv\n"
            "    key_from: employee_id\n    columns:\n"
            "      - column: Hire Date\n        source: hire_date\n",
        )
        with pytest.raises(GenerateError, match=r"required column.*Legal Name"):
            generate_workbook(template, spec, base / "f.xlsx")

    def test_mapping_the_key_column_is_rejected(self, tmp_path: Path) -> None:
        template, base = self._base(tmp_path)
        spec = self._spec(
            base,
            "sheets:\n  - sheet: Workers\n    source: workers.csv\n"
            "    key_from: employee_id\n    columns:\n"
            "      - column: Spreadsheet Key\n        source: employee_id\n"
            "      - column: Legal Name\n        source: name\n",
        )
        with pytest.raises(GenerateError, match="assigned automatically"):
            generate_workbook(template, spec, base / "f.xlsx")

    def test_missing_key_from_on_keyed_sheet(self, tmp_path: Path) -> None:
        template, base = self._base(tmp_path)
        spec = self._spec(
            base,
            "sheets:\n  - sheet: Workers\n    source: workers.csv\n    columns:\n"
            "      - column: Legal Name\n        source: name\n",
        )
        with pytest.raises(GenerateError, match="key_from"):
            generate_workbook(template, spec, base / "f.xlsx")

    def test_missing_csv_field(self, tmp_path: Path) -> None:
        template, base = self._base(tmp_path)
        spec = self._spec(
            base,
            "sheets:\n  - sheet: Workers\n    source: workers.csv\n"
            "    key_from: employee_id\n    columns:\n"
            "      - column: Legal Name\n        source: full_name\n",
        )
        with pytest.raises(GenerateError, match=r"no field.*full_name"):
            generate_workbook(template, spec, base / "f.xlsx")

    def test_missing_csv_file(self, tmp_path: Path) -> None:
        template, base = self._base(tmp_path)
        spec = self._spec(
            base,
            "sheets:\n  - sheet: Workers\n    source: nowhere.csv\n"
            "    key_from: employee_id\n    columns:\n"
            "      - column: Legal Name\n        source: name\n",
        )
        with pytest.raises(GenerateError, match="not found"):
            generate_workbook(template, spec, base / "f.xlsx")

    def test_source_and_const_together_rejected(self) -> None:
        with pytest.raises(GenerateError, match="exactly one"):
            ColumnMapping(column="X", source="a", const="b")
        with pytest.raises(GenerateError, match="exactly one"):
            ColumnMapping(column="X")

    def test_unknown_key_strategy_rejected(self) -> None:
        with pytest.raises(GenerateError, match="key_strategy"):
            LoadSpec(sheets=[SheetMapping(sheet="S", source="s.csv")], key_strategy="chaos")


class TestSpecFiles:
    def test_json_spec_loads(self, tmp_path: Path) -> None:
        path = tmp_path / "load.json"
        path.write_text(
            '{"sheets": [{"sheet": "Workers", "source": "workers.csv",'
            ' "key_from": "employee_id",'
            ' "columns": [{"column": "Legal Name", "source": "name"}]}]}',
            encoding="utf-8",
        )
        spec = load_spec(path)
        assert spec.key_strategy == "sequential"
        assert spec.sheets[0].key_from == "employee_id"
        assert spec.base_dir == str(tmp_path)

    def test_unknown_spec_keys_rejected(self, tmp_path: Path) -> None:
        path = tmp_path / "load.yaml"
        path.write_text(
            "sheets:\n  - sheet: A\n    source: a.csv\n    tenant: prod\n",
            encoding="utf-8",
        )
        with pytest.raises(GenerateError, match=r"unknown key.*tenant"):
            load_spec(path)

    def test_empty_spec_rejected(self, tmp_path: Path) -> None:
        path = tmp_path / "load.yaml"
        path.write_text("sheets: []\n", encoding="utf-8")
        with pytest.raises(GenerateError, match="maps no sheets"):
            load_spec(path)
