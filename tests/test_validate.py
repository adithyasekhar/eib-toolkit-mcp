"""Tests for the rule-based validation engine — one section per rule family.

Most cases round-trip through a real .xlsx (written with openpyxl, parsed by
the tolerant parser) so validation is exercised exactly as users hit it. The
XML-unsafe-character cases build model objects in memory instead, because
openpyxl itself (correctly) refuses to write control characters to a file.
"""

import datetime as dt
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook as XlsxWorkbook

from eib_toolkit.model import (
    ColumnSpec,
    ColumnType,
    SheetData,
    SheetRole,
    SheetSpec,
    TemplateSpec,
    Workbook,
)
from eib_toolkit.parser import parse_workbook
from eib_toolkit.validate import (
    RULES,
    Severity,
    ValidationConfig,
    validate_workbook,
)

WID = "d588c41a446c11de98360015c5e6daf6"  # 32 hex chars, synthetic


def _write(path: Path, sheets: dict[str, list[list[Any]]]) -> Path:
    wb = XlsxWorkbook()
    assert wb.active is not None
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    wb.save(path)
    return path


_HEADER = [
    "Spreadsheet Key*",
    "Employee_Reference",
    "Legal Name",
    "Hire Date",
    "Annual Amount",
    "Active",
]
_HINTS = [
    "Numeric. Required.",
    "Reference ID Type: Employee_ID",
    "Text. Required.",
    "Date (yyyy-mm-dd)",
    "Numeric. Optional.",
    "Boolean (Y/N)",
]


def _load(
    tmp_path: Path,
    rows: list[list[Any]],
    plans: list[list[Any]] | None = None,
    config: ValidationConfig | None = None,
):
    """Write a workbook with the standard header band + ``rows``, validate it."""
    sheets: dict[str, list[list[Any]]] = {"Submit Data": [_HEADER, _HINTS, *rows]}
    if plans is not None:
        sheets["Plan Assignments"] = [
            ["Spreadsheet Key*", "Plan Name"],
            ["Numeric. Required.", "Text. Required."],
            *plans,
        ]
    wb = parse_workbook(_write(tmp_path / "load.xlsx", sheets))
    return validate_workbook(wb, config)


def _clean_row(key: int = 1) -> list[Any]:
    return [key, "EMP-001", "Avery Example", dt.date(2026, 1, 15), 55000, "Y"]


def _codes(report) -> list[str]:
    return [f.code for f in report.findings]


class TestCleanLoad:
    def test_no_findings_and_ok(self, tmp_path: Path) -> None:
        report = _load(tmp_path, [_clean_row(1), _clean_row(2)], plans=[[1, "Medical"]])
        assert report.findings == []
        assert report.ok

    def test_iso_date_text_and_typed_values_pass(self, tmp_path: Path) -> None:
        row = [1, "EMP-001", "Avery Example", "2026-01-15", 55000.5, True]
        assert _codes(_load(tmp_path, [row])) == []

    def test_interior_blank_row_is_skipped(self, tmp_path: Path) -> None:
        rows = [_clean_row(1), [None] * 6, _clean_row(2)]
        assert _codes(_load(tmp_path, rows)) == []


class TestRequiredCells:
    def test_blank_required_cell_is_error(self, tmp_path: Path) -> None:
        row = [1, "EMP-001", None, dt.date(2026, 1, 15), None, "Y"]
        report = _load(tmp_path, [row])
        assert _codes(report) == ["EIB001"]  # Annual Amount is optional: no finding
        f = report.findings[0]
        assert f.severity is Severity.ERROR
        assert f.location == "Submit Data!C3"
        assert "Legal Name" in f.message
        assert not report.ok

    def test_whitespace_only_counts_as_blank(self, tmp_path: Path) -> None:
        row = [1, "EMP-001", "   ", dt.date(2026, 1, 15), None, "Y"]
        assert _codes(_load(tmp_path, [row])) == ["EIB001"]


class TestDates:
    def test_unparseable_date_is_error(self, tmp_path: Path) -> None:
        row = [1, "EMP-001", "Avery Example", "next monday", None, "Y"]
        report = _load(tmp_path, [row])
        assert _codes(report) == ["EIB010"]
        assert report.findings[0].evidence == "'next monday'"

    @pytest.mark.parametrize(
        "text", ["01/15/2026", "15.01.2026", "15-Jan-2026", "Jan 15, 2026", "2026/01/15"]
    )
    def test_locale_date_is_warning(self, tmp_path: Path, text: str) -> None:
        row = [1, "EMP-001", "Avery Example", text, None, "Y"]
        report = _load(tmp_path, [row])
        assert _codes(report) == ["EIB011"]
        assert report.findings[0].severity is Severity.WARNING
        assert report.ok  # warnings alone leave the load loadable


class TestNumbers:
    def test_non_numeric_text_is_error(self, tmp_path: Path) -> None:
        row = [1, "EMP-001", "Avery Example", dt.date(2026, 1, 15), "sixty grand", "Y"]
        assert _codes(_load(tmp_path, [row])) == ["EIB020"]

    def test_boolean_in_numeric_column_is_error(self, tmp_path: Path) -> None:
        row = [1, "EMP-001", "Avery Example", dt.date(2026, 1, 15), True, "Y"]
        assert _codes(_load(tmp_path, [row])) == ["EIB020"]

    @pytest.mark.parametrize(
        ("text", "canonical"),
        [
            ("55,000.00", "55000.00"),
            ("1.234.567,89", "1234567.89"),
            ("1234,56", "1234.56"),
            ("61000", "61000"),  # plain number stored as text still needs coercion
        ],
    )
    def test_coercible_numeric_text_is_warning(
        self, tmp_path: Path, text: str, canonical: str
    ) -> None:
        row = [1, "EMP-001", "Avery Example", dt.date(2026, 1, 15), text, "Y"]
        report = _load(tmp_path, [row])
        assert _codes(report) == ["EIB021"]
        assert canonical in report.findings[0].message


class TestBooleans:
    @pytest.mark.parametrize("value", ["Y", "n", "Yes", "NO", "TRUE", "false", 1, 0])
    def test_recognized_tokens_pass(self, tmp_path: Path, value: Any) -> None:
        row = [1, "EMP-001", "Avery Example", dt.date(2026, 1, 15), None, value]
        assert _codes(_load(tmp_path, [row])) == []

    def test_unrecognized_token_is_error(self, tmp_path: Path) -> None:
        row = [1, "EMP-001", "Avery Example", dt.date(2026, 1, 15), None, "maybe"]
        assert _codes(_load(tmp_path, [row])) == ["EIB030"]


class TestReferenceIds:
    def test_wid_in_named_reference_column_is_warning(self, tmp_path: Path) -> None:
        row = [1, WID, "Avery Example", dt.date(2026, 1, 15), None, "Y"]
        report = _load(tmp_path, [row])
        assert _codes(report) == ["EIB040"]
        assert "Employee_ID" in report.findings[0].message

    def test_named_id_passes(self, tmp_path: Path) -> None:
        assert _codes(_load(tmp_path, [_clean_row()])) == []

    def test_non_wid_in_wid_column_is_error(self, tmp_path: Path) -> None:
        sheets = {
            "Data": [
                ["Spreadsheet Key*", "Worker_Reference"],
                ["Numeric. Required.", "Reference ID Type: WID"],
                [1, "EMP-001"],
                [2, WID],
            ]
        }
        wb = parse_workbook(_write(tmp_path / "t.xlsx", sheets))
        report = validate_workbook(wb)
        assert _codes(report) == ["EIB041"]
        assert report.findings[0].location == "Data!B3"


class TestSpreadsheetKeys:
    def test_duplicate_primary_key_is_error(self, tmp_path: Path) -> None:
        report = _load(tmp_path, [_clean_row(1), _clean_row(2), _clean_row(1)])
        assert _codes(report) == ["EIB050"]
        f = report.findings[0]
        assert f.location == "Submit Data!A5"  # the *second* occurrence is flagged
        assert "row 3" in f.message

    def test_missing_key_on_repeating_row_is_error(self, tmp_path: Path) -> None:
        report = _load(tmp_path, [_clean_row(1)], plans=[[None, "Medical"]])
        codes = _codes(report)
        assert "EIB051" in codes
        # The blank key is also a blank required cell; both findings coexist.
        assert codes.count("EIB001") == 1

    def test_orphaned_key_is_error(self, tmp_path: Path) -> None:
        report = _load(tmp_path, [_clean_row(1)], plans=[[1, "Medical"], [9, "Dental"]])
        assert _codes(report) == ["EIB052"]
        assert report.findings[0].location == "Plan Assignments!A4"

    def test_key_join_normalizes_numeric_text_and_floats(self, tmp_path: Path) -> None:
        # Primary declares key 1 as an int; children reference it as "1" and 1.0.
        report = _load(tmp_path, [_clean_row(1)], plans=[["1", "Medical"], [1.0, "Dental"]])
        # The text key "1" still (correctly) draws a stored-as-text warning,
        # but the join itself resolves: no duplicate/missing/orphan findings.
        assert _codes(report) == ["EIB021"]
        assert not any(f.code in {"EIB050", "EIB051", "EIB052"} for f in report.findings)


class TestCeilings:
    def test_row_ceiling_is_sheet_level_warning(self, tmp_path: Path) -> None:
        config = ValidationConfig(max_rows_per_sheet=2)
        report = _load(tmp_path, [_clean_row(1), _clean_row(2), _clean_row(3)], config=config)
        assert _codes(report) == ["EIB060"]
        f = report.findings[0]
        assert (f.row, f.column) == (0, 0)
        assert f.location == "Submit Data"

    def test_cell_length_ceiling_is_error(self, tmp_path: Path) -> None:
        config = ValidationConfig(max_cell_length=20)
        row = [1, "EMP-001", "X" * 21, dt.date(2026, 1, 15), None, "Y"]
        report = _load(tmp_path, [row], config=config)
        assert _codes(report) == ["EIB061"]
        assert "21 characters" in report.findings[0].message


class TestWhitespaceAndHeaderlessColumns:
    def test_leading_trailing_whitespace_is_warning(self, tmp_path: Path) -> None:
        row = [1, " EMP-001 ", "Avery Example", dt.date(2026, 1, 15), None, "Y"]
        report = _load(tmp_path, [row])
        assert _codes(report) == ["EIB071"]
        assert report.findings[0].evidence == "' EMP-001 '"

    def test_value_in_headerless_column_is_warning(self, tmp_path: Path) -> None:
        rows = [[*_clean_row(1), "stray note"]]  # one column past the declared band
        report = _load(tmp_path, rows)
        assert _codes(report) == ["EIB080"]
        assert report.findings[0].location == "Submit Data!G3"


def _in_memory_sheet(rows: list[list[Any]]) -> Workbook:
    """A minimal one-sheet Workbook built without touching disk."""
    spec = SheetSpec(
        name="Data",
        role=SheetRole.PRIMARY,
        header_row=1,
        data_start_row=2,
        columns=[
            ColumnSpec(index=1, header="Note", col_type=ColumnType.TEXT, required=True),
        ],
    )
    template = TemplateSpec(source="<memory>", sheets=[spec])
    data = SheetData(spec=spec, rows=rows)
    return Workbook(path="<memory>", template=template, data=[data])


class TestXmlUnsafeCharacters:
    @pytest.mark.parametrize("bad", ["\x00", "\x0b", "\x1f", "﷐", "￿"])
    def test_unsafe_character_is_error(self, bad: str) -> None:
        report = validate_workbook(_in_memory_sheet([[f"note{bad}text"]]))
        assert _codes(report) == ["EIB070"]
        assert "U+" in report.findings[0].message

    def test_tabs_newlines_and_unicode_text_pass(self) -> None:
        report = validate_workbook(_in_memory_sheet([["line one\nline\ttwo — üñïçødé"]]))
        assert _codes(report) == []

    def test_each_distinct_character_reported_once(self) -> None:
        report = validate_workbook(_in_memory_sheet([["a\x00b\x00c\x01d"]]))
        assert _codes(report) == ["EIB070", "EIB070"]


class TestReportShape:
    def test_findings_sorted_in_reading_order(self, tmp_path: Path) -> None:
        rows = [
            [1, "EMP-001", None, "bogus", None, "Y"],  # C3 blank required, D3 bad date
            [1, "EMP-002", "Blake Sample", dt.date(2026, 2, 1), None, "maybe"],  # A4 dup, F4
        ]
        report = _load(tmp_path, rows, plans=[[9, "Medical"]])
        locations = [f.location for f in report.findings]
        assert locations == [
            "Submit Data!C3",
            "Submit Data!D3",
            "Submit Data!A4",
            "Submit Data!F4",
            "Plan Assignments!A3",
        ]

    def test_rollups_and_to_dict(self, tmp_path: Path) -> None:
        rows = [[1, "EMP-001", None, "01/15/2026", None, "Y"]]
        report = _load(tmp_path, rows)
        assert report.by_severity() == {"error": 1, "warning": 1, "info": 0}
        assert report.by_code() == {"EIB001": 1, "EIB011": 1}
        assert report.by_sheet() == {"Submit Data": 2}

        d = report.to_dict()
        assert d["ok"] is False
        assert d["summary"]["total"] == 2
        assert d["findings"][0]["location"] == "Submit Data!C3"
        assert d["findings"][0]["severity"] == "error"

    def test_per_rule_cap_adds_suppression_note(self, tmp_path: Path) -> None:
        config = ValidationConfig(max_findings_per_rule_per_sheet=3)
        rows = [[k, "EMP-001", None, dt.date(2026, 1, 15), None, "Y"] for k in range(1, 7)]
        report = _load(tmp_path, rows, config=config)
        cell_findings = [f for f in report.findings if f.code == "EIB001" and f.row]
        notes = [f for f in report.findings if f.code == "EIB001" and not f.row]
        assert len(cell_findings) == 3
        assert len(notes) == 1
        assert "3 further finding(s)" in notes[0].message

    def test_every_emitted_code_is_registered(self) -> None:
        assert all(code.startswith("EIB") and len(code) == 6 for code in RULES)
        severities = {sev for sev, _ in RULES.values()}
        assert severities <= set(Severity)
