"""Generate filled EIB load workbooks from a declarative load spec + CSVs.

The generator's contract: take a Workday-generated EIB template, a small
declarative *load spec* (YAML or JSON), and one input CSV per target sheet,
and produce a filled .xlsx that loads cleanly. It fills a **copy of the
template file itself**, so the multi-row header band the EIB transform keys
on is preserved byte-for-byte rather than approximated. Spreadsheet keys for
multi-sheet loads are auto-assigned: primary rows get sequential keys, and
repeating-group rows join back to them through a shared business-key CSV
column — nobody should be hand-numbering join keys in Excel.

Structural problems (missing sheet, unmapped required column, unknown CSV
field) raise :class:`GenerateError` before anything is written. Value-level
problems (an uncoercible date, an unknown business key) never silently
mutate data: the raw value is written as-is and reported as a warning, so
``validate_workbook`` on the output shows exactly what a loader would see.

Load spec shape (YAML shown; JSON is the same structure)::

    sheets:
      - sheet: Worker Data            # template sheet name
        source: workers.csv           # CSV path, relative to the spec file
        key_from: employee_id         # business-key CSV column (multi-sheet loads)
        columns:
          - column: Employee_Reference  # template column header
            source: employee_id         # CSV field ...
          - column: Currency
            const: USD                  # ... or a constant, never both
"""

from __future__ import annotations

import csv
import datetime as dt
import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from eib_toolkit.model import ColumnSpec, ColumnType, SheetSpec, TemplateSpec
from eib_toolkit.parser import parse_template

__all__ = [
    "ColumnMapping",
    "GenerateError",
    "GenerateReport",
    "LoadSpec",
    "SheetMapping",
    "generate_workbook",
    "load_spec",
    "spec_from_dict",
]


class GenerateError(ValueError):
    """A load spec cannot be applied to the template it targets."""


# ---------------------------------------------------------------------------
# Load-spec model


@dataclass
class ColumnMapping:
    """Map one template column from a CSV field or a constant (exactly one)."""

    column: str
    source: str = ""
    const: Any = None

    def __post_init__(self) -> None:
        if bool(self.source) == (self.const is not None):
            raise GenerateError(
                f"column {self.column!r}: set exactly one of 'source' or 'const'"
            )


@dataclass
class SheetMapping:
    """One target sheet: where its rows come from and how columns map."""

    sheet: str
    source: str  # CSV path, resolved relative to the spec file
    columns: list[ColumnMapping] = field(default_factory=list)
    key_from: str = ""  # business-key CSV column used to join multi-sheet loads


@dataclass
class LoadSpec:
    """A whole load: every sheet mapping plus the key strategy.

    ``key_strategy`` is ``"sequential"`` (default: primary rows are numbered
    1..N in input order and repeating rows join via ``key_from``) or
    ``"from_source"`` (the ``key_from`` CSV value **is** the spreadsheet key,
    written verbatim on every sheet).
    """

    sheets: list[SheetMapping] = field(default_factory=list)
    key_strategy: str = "sequential"
    base_dir: str = "."  # directory CSV paths resolve against

    def __post_init__(self) -> None:
        if self.key_strategy not in ("sequential", "from_source"):
            raise GenerateError(
                f"unknown key_strategy {self.key_strategy!r}; "
                "expected 'sequential' or 'from_source'"
            )
        if not self.sheets:
            raise GenerateError("load spec maps no sheets")


def load_spec(path: str | Path) -> LoadSpec:
    """Read a YAML or JSON load spec file into a :class:`LoadSpec`."""
    path = Path(path)
    text = path.read_text(encoding="utf-8")
    if path.suffix.lower() == ".json":
        raw = json.loads(text)
    else:
        import yaml  # deferred so JSON-only users never need it

        raw = yaml.safe_load(text)
    return spec_from_dict(raw, base_dir=str(path.parent), origin=str(path))


def spec_from_dict(raw: Any, base_dir: str = ".", origin: str = "load spec") -> LoadSpec:
    """Build a :class:`LoadSpec` from already-parsed YAML/JSON data.

    This is the single structural gate every spec passes through — files via
    :func:`load_spec`, drafted specs (e.g. from the optional Claude layer)
    directly. ``origin`` labels error messages.
    """
    if not isinstance(raw, dict):
        raise GenerateError(f"{origin}: load spec must be a mapping, got {type(raw).__name__}")

    unknown = set(raw) - {"sheets", "key_strategy"}
    if unknown:
        raise GenerateError(f"{origin}: unknown load-spec key(s): {', '.join(sorted(unknown))}")

    sheets = []
    for i, entry in enumerate(raw.get("sheets") or []):
        if not isinstance(entry, dict):
            raise GenerateError(f"{origin}: sheets[{i}] must be a mapping")
        unknown = set(entry) - {"sheet", "source", "columns", "key_from"}
        if unknown:
            raise GenerateError(
                f"{origin}: sheets[{i}]: unknown key(s): {', '.join(sorted(unknown))}"
            )
        for want in ("sheet", "source"):
            if not entry.get(want):
                raise GenerateError(f"{origin}: sheets[{i}] is missing {want!r}")
        columns = []
        for j, col in enumerate(entry.get("columns") or []):
            if not isinstance(col, dict) or not col.get("column"):
                raise GenerateError(f"{origin}: sheets[{i}].columns[{j}] needs a 'column'")
            unknown = set(col) - {"column", "source", "const"}
            if unknown:
                raise GenerateError(
                    f"{origin}: sheets[{i}].columns[{j}]: unknown key(s): "
                    f"{', '.join(sorted(unknown))}"
                )
            columns.append(
                ColumnMapping(
                    column=str(col["column"]),
                    source=str(col.get("source") or ""),
                    const=col.get("const"),
                )
            )
        sheets.append(
            SheetMapping(
                sheet=str(entry["sheet"]),
                source=str(entry["source"]),
                columns=columns,
                key_from=str(entry.get("key_from") or ""),
            )
        )
    return LoadSpec(
        sheets=sheets,
        key_strategy=str(raw.get("key_strategy") or "sequential"),
        base_dir=base_dir,
    )


# ---------------------------------------------------------------------------
# Generation report


@dataclass
class GenerateReport:
    """What was written where, plus every value-level warning."""

    template: str
    output: str
    rows_written: dict[str, int] = field(default_factory=dict)
    key_assignments: dict[str, int] = field(default_factory=dict)  # business key -> sheet key
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


# ---------------------------------------------------------------------------
# The generator


def generate_workbook(
    template_path: str | Path,
    spec: LoadSpec,
    output_path: str | Path,
) -> GenerateReport:
    """Fill a copy of ``template_path`` per ``spec`` and save to ``output_path``."""
    template_path = Path(template_path)
    output_path = Path(output_path)
    template = parse_template(template_path)

    plans = [_plan_sheet(template, mapping, spec) for mapping in spec.sheets]
    report = GenerateReport(template=str(template_path), output=str(output_path))

    # Primary sheet first: its pass fixes the business-key -> spreadsheet-key map
    # that every repeating-group sheet joins through.
    primary_name = _primary_name(template, spec)
    plans.sort(key=lambda p: p.sheet_spec.name != primary_name)

    wb = load_workbook(template_path)  # a copy of the template, header band intact
    try:
        keymap: dict[str, int] = {}
        for plan in plans:
            is_primary = plan.sheet_spec.name == primary_name
            _fill_sheet(wb[plan.sheet_spec.name], plan, spec, keymap, is_primary, report)
        report.key_assignments = dict(keymap)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
    finally:
        wb.close()
    return report


def _primary_name(template: TemplateSpec, spec: LoadSpec) -> str:
    primary = template.primary_sheet()
    if primary is not None and any(m.sheet == primary.name for m in spec.sheets):
        return primary.name
    return spec.sheets[0].sheet


@dataclass
class _SheetPlan:
    mapping: SheetMapping
    sheet_spec: SheetSpec
    cells: list[tuple[ColumnSpec, ColumnMapping]]  # resolved target column per mapping
    rows: list[dict[str, str]]  # CSV rows


def _plan_sheet(template: TemplateSpec, mapping: SheetMapping, spec: LoadSpec) -> _SheetPlan:
    """Resolve one sheet mapping against the template; all-or-nothing."""
    sheet_spec = template.sheet(mapping.sheet)
    if sheet_spec is None or not sheet_spec.columns:
        known = ", ".join(s.name for s in template.data_sheets())
        raise GenerateError(
            f"template has no data sheet named {mapping.sheet!r} (data sheets: {known})"
        )

    cells = []
    for col_map in mapping.columns:
        target = sheet_spec.column_by_header(col_map.column)
        if target is None:
            raise GenerateError(
                f"sheet {mapping.sheet!r} has no column {col_map.column!r} "
                f"(headers: {', '.join(c.header for c in sheet_spec.columns if c.header)})"
            )
        if target.is_key:
            raise GenerateError(
                f"sheet {mapping.sheet!r}: column {target.header!r} is the spreadsheet "
                "key; it is assigned automatically and cannot be mapped"
            )
        cells.append((target, col_map))

    seen: set[int] = set()
    for target, _ in cells:
        if target.index in seen:
            raise GenerateError(
                f"sheet {mapping.sheet!r}: column {target.header!r} is mapped twice"
            )
        seen.add(target.index)

    unmapped_required = [
        c.header
        for c in sheet_spec.columns
        if c.required and not c.is_key and c.index not in seen
    ]
    if unmapped_required:
        raise GenerateError(
            f"sheet {mapping.sheet!r}: required column(s) not mapped: "
            f"{', '.join(unmapped_required)}"
        )

    if sheet_spec.key_column is not None and not mapping.key_from:
        raise GenerateError(
            f"sheet {mapping.sheet!r} has a spreadsheet-key column; the mapping "
            "needs 'key_from' naming the business-key CSV field"
        )

    rows = _read_csv(Path(spec.base_dir) / mapping.source)
    fields = set(rows[0]) if rows else set()
    missing = sorted(
        {m.source for m in mapping.columns if m.source}
        | ({mapping.key_from} if mapping.key_from and rows else set())
    )
    missing = [f for f in missing if rows and f not in fields]
    if missing:
        raise GenerateError(
            f"CSV {mapping.source!r} has no field(s): {', '.join(missing)} "
            f"(fields: {', '.join(sorted(fields))})"
        )
    return _SheetPlan(mapping=mapping, sheet_spec=sheet_spec, cells=cells, rows=rows)


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        raise GenerateError(f"input CSV not found: {path}")
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        if not reader.fieldnames:
            raise GenerateError(f"input CSV has no header row: {path}")
        return [dict(row) for row in reader]


def _fill_sheet(
    ws: Any,
    plan: _SheetPlan,
    spec: LoadSpec,
    keymap: dict[str, int],
    is_primary: bool,
    report: GenerateReport,
) -> None:
    sheet_spec = plan.sheet_spec
    key_col = sheet_spec.key_column
    row_cursor = sheet_spec.data_start_row or (sheet_spec.header_row + 1)

    for i, csv_row in enumerate(plan.rows):
        if key_col is not None:
            key_value = _resolve_key(plan, csv_row, i, spec, keymap, is_primary, report)
            if key_value is not None:
                ws.cell(row=row_cursor, column=key_col.index, value=key_value)
        for target, col_map in plan.cells:
            raw = col_map.const if col_map.const is not None else csv_row.get(col_map.source, "")
            value, warning = _coerce(raw, target)
            if warning:
                report.warnings.append(
                    f"{sheet_spec.name}!{target.letter}{row_cursor}: {warning}"
                )
            if value is not None:
                ws.cell(row=row_cursor, column=target.index, value=value)
        row_cursor += 1

    report.rows_written[sheet_spec.name] = len(plan.rows)


def _resolve_key(
    plan: _SheetPlan,
    csv_row: dict[str, str],
    row_index: int,
    spec: LoadSpec,
    keymap: dict[str, int],
    is_primary: bool,
    report: GenerateReport,
) -> int | str | None:
    business = (csv_row.get(plan.mapping.key_from) or "").strip()
    if not business:
        report.warnings.append(
            f"{plan.sheet_spec.name}: CSV row {row_index + 2} has a blank "
            f"{plan.mapping.key_from!r}; spreadsheet key left empty"
        )
        return None
    if spec.key_strategy == "from_source":
        return business

    token = business.casefold()
    if is_primary:
        if token in keymap:
            report.warnings.append(
                f"{plan.sheet_spec.name}: duplicate business key {business!r} on the "
                f"primary sheet (CSV row {row_index + 2}); rows share spreadsheet key "
                f"{keymap[token]}"
            )
        else:
            keymap[token] = len(keymap) + 1
        return keymap[token]
    if token not in keymap:
        report.warnings.append(
            f"{plan.sheet_spec.name}: business key {business!r} (CSV row {row_index + 2}) "
            "has no matching primary row; spreadsheet key left empty"
        )
        return None
    return keymap[token]


# ---------------------------------------------------------------------------
# Type coercion — writes real .xlsx types so Workday's transform sees what a
# careful human would have entered. Never lossy: uncoercible input is written
# verbatim and warned about, for the validator to flag with full evidence.

_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_TRUE_TOKENS = frozenset({"y", "yes", "true", "1"})
_FALSE_TOKENS = frozenset({"n", "no", "false", "0"})


def _coerce(raw: Any, target: ColumnSpec) -> tuple[Any, str]:
    """(value to write, warning) — warning is "" when the value is clean."""
    if raw is None:
        return None, ""
    if isinstance(raw, (int, float, bool, dt.date)):
        return raw, ""  # constants may already be typed
    text = str(raw).strip()
    if text == "":
        return None, ""

    if target.col_type is ColumnType.DATE:
        if _ISO_DATE_RE.match(text):
            return dt.date.fromisoformat(text), ""
        return text, (
            f"{text!r} is not an ISO 8601 date (yyyy-mm-dd) for date column "
            f"{target.header!r}; written as-is"
        )
    if target.col_type is ColumnType.NUMERIC:
        try:
            return int(text), ""
        except ValueError:
            pass
        try:
            return float(text), ""
        except ValueError:
            return text, (
                f"{text!r} is not numeric for column {target.header!r}; written as-is"
            )
    if target.col_type is ColumnType.BOOLEAN:
        low = text.casefold()
        if low in _TRUE_TOKENS:
            return "Y", ""
        if low in _FALSE_TOKENS:
            return "N", ""
        return text, (
            f"{text!r} is not a recognizable boolean for column {target.header!r}; "
            "written as-is"
        )
    return text, ""
