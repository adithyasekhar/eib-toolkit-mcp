"""Tolerant parser for EIB templates and filled load workbooks.

Workday's generated EIB .xlsx files are structured but messy in practice:
multi-row header bands whose exact depth varies, merged group-header cells,
stray instruction/overview sheets, blank padding rows and columns, and hint
rows whose wording differs between operations and releases. This parser
recovers a :class:`~eib_toolkit.model.TemplateSpec` (and, for filled
workbooks, the data) using deterministic heuristics rather than assuming any
fixed layout — parse tolerantly, report precisely.
"""

from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from eib_toolkit.model import (
    ColumnSpec,
    ColumnType,
    SheetData,
    SheetRole,
    SheetSpec,
    TemplateSpec,
    Workbook,
)

# How many leading rows of a sheet are scanned for the header band.
_SCAN_ROWS = 12
# At most this many hint rows are consumed below the field-name row.
_MAX_HINT_ROWS = 4
# Data-type inference looks at up to this many non-blank cells per column.
_INFER_SAMPLE = 50

_INSTRUCTION_SHEET_RE = re.compile(r"instruction|overview|readme|help|about", re.IGNORECASE)
_SPREADSHEET_KEY_RE = re.compile(r"spreadsheet\s*key", re.IGNORECASE)

# Tokens that mark a row as a hint row (type/format/required annotations).
_HINT_ROW_RE = re.compile(
    r"required|optional|reference\s*id|format|yyyy|y/n|true/false"
    r"|\btext\b|\bdate\b|\bboolean\b|\bnumeric\b|\bnumber\b|\bdecimal\b|\binteger\b",
    re.IGNORECASE,
)

# Declared-type detection, checked in order (first match wins).
_TYPE_PATTERNS: list[tuple[ColumnType, re.Pattern[str]]] = [
    (ColumnType.REFERENCE, re.compile(r"reference\s*id|_reference\b", re.IGNORECASE)),
    (ColumnType.DATE, re.compile(r"\bdate(?:time)?\b|yyyy-mm-dd", re.IGNORECASE)),
    (ColumnType.BOOLEAN, re.compile(r"\bbool(?:ean)?\b|\by/n\b|true/false", re.IGNORECASE)),
    (
        ColumnType.NUMERIC,
        re.compile(r"\bnumeric\b|\bnumber\b|\bdecimal\b|\binteger\b|\bamount\b|\bcurrency\b",
                   re.IGNORECASE),
    ),
    (ColumnType.TEXT, re.compile(r"\btext\b|\bstring\b", re.IGNORECASE)),
]

_REF_ID_TYPE_RE = re.compile(
    r"reference\s*id(?:\s*type)?\s*[:=]\s*([A-Za-z0-9_]+)", re.IGNORECASE
)

_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}([T ]\d{2}:\d{2}(:\d{2})?)?$")
_NUMERIC_TEXT_RE = re.compile(r"^-?\d{1,3}(?:,\d{3})*(?:\.\d+)?$|^-?\d+(?:[.,]\d+)?$")
_BOOL_TOKENS = frozenset({"y", "n", "yes", "no", "true", "false"})


def parse_workbook(path: str | Path) -> Workbook:
    """Parse an EIB workbook (template or filled load) into structure + data."""
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        spec = TemplateSpec(source=str(path))
        data: list[SheetData] = []
        for ws in wb.worksheets:
            grid = _trim_grid([list(row) for row in ws.iter_rows(values_only=True)])
            sheet_spec, rows = _parse_sheet(ws.title, grid)
            spec.sheets.append(sheet_spec)
            if sheet_spec.role in (SheetRole.PRIMARY, SheetRole.REPEATING):
                data.append(SheetData(spec=sheet_spec, rows=rows))
        _assign_roles(spec)
        return Workbook(path=str(path), template=spec, data=data)
    finally:
        wb.close()


def parse_template(path: str | Path) -> TemplateSpec:
    """Parse just the structure of an EIB workbook (data rows ignored)."""
    return parse_workbook(path).template


# ---------------------------------------------------------------------------
# Sheet-level parsing


def _parse_sheet(name: str, grid: list[list[Any]]) -> tuple[SheetSpec, list[list[Any]]]:
    if not grid:
        return SheetSpec(name=name, role=SheetRole.EMPTY), []

    header_idx = _find_header_row(grid)  # 0-based index into grid, or None
    if header_idx is None or _INSTRUCTION_SHEET_RE.search(name):
        return SheetSpec(name=name, role=SheetRole.INSTRUCTIONS), []

    header = [_text(v) for v in grid[header_idx]]
    group_row = _group_row(grid, header_idx)
    hint_rows = _hint_rows(grid, header_idx)
    data_idx = header_idx + 1 + len(hint_rows)

    # Skip blank padding rows between the header band and the data region.
    while data_idx < len(grid) and _is_blank_row(grid[data_idx]):
        data_idx += 1
    rows = _trim_trailing_blank_rows(grid[data_idx:])

    n_cols = max(
        [len(header)]
        + [_used_width(r) for r in rows]
        + [_used_width(h) for h in hint_rows]
    )
    columns = [
        _build_column(
            index=c + 1,
            header=header[c] if c < len(header) else "",
            group=group_row[c] if c < len(group_row) else "",
            hints=[h[c] for h in hint_rows if c < len(h) and _text(h[c])],
            sample=[r[c] for r in rows if c < len(r)],
        )
        for c in range(n_cols)
    ]

    sheet_spec = SheetSpec(
        name=name,
        role=SheetRole.PRIMARY,  # provisional; _assign_roles finalizes
        header_row=header_idx + 1,
        data_start_row=data_idx + 1,
        columns=columns,
    )
    return sheet_spec, rows


def _find_header_row(grid: list[list[Any]]) -> int | None:
    """Pick the field-name row: the earliest all-string row with the most
    distinct non-blank cells within the scan window."""
    best_idx: int | None = None
    best_score = 0
    for i, row in enumerate(grid[:_SCAN_ROWS]):
        cells = [v for v in row if v is not None and _text(v)]
        if not cells or not all(isinstance(v, str) for v in cells):
            continue  # typed cells mean data, not header
        score = len({_text(v).casefold() for v in cells})
        if score > best_score:
            best_idx, best_score = i, score
    # EIB data sheets always carry several columns (a spreadsheet key plus
    # fields); a "best row" of one distinct cell is prose, not a header band.
    if best_idx is None or best_score < 2:
        return None
    return best_idx


def _group_row(grid: list[list[Any]], header_idx: int) -> list[str]:
    """Nearest non-blank row above the header, forward-filled across the
    blanks that merged group cells leave behind."""
    for i in range(header_idx - 1, -1, -1):
        if not _is_blank_row(grid[i]):
            filled: list[str] = []
            last = ""
            for v in grid[i]:
                text = _text(v)
                if text:
                    last = text
                filled.append(last)
            return filled
    return []


def _hint_rows(grid: list[list[Any]], header_idx: int) -> list[list[Any]]:
    hints: list[list[Any]] = []
    for row in grid[header_idx + 1 : header_idx + 1 + _MAX_HINT_ROWS]:
        cells = [v for v in row if v is not None and _text(v)]
        if not cells or not all(isinstance(v, str) for v in cells):
            break
        if not any(_HINT_ROW_RE.search(v) for v in cells):
            break
        hints.append(row)
    return hints


def _assign_roles(spec: TemplateSpec) -> None:
    """First data sheet is primary; later keyed sheets are repeating groups."""
    seen_primary = False
    for sheet in spec.sheets:
        if sheet.role is not SheetRole.PRIMARY:
            continue
        if not seen_primary:
            seen_primary = True
        elif sheet.key_column is not None:
            sheet.role = SheetRole.REPEATING


# ---------------------------------------------------------------------------
# Column classification


def _build_column(
    index: int, header: str, group: str, hints: list[str], sample: list[Any]
) -> ColumnSpec:
    header = header.strip()
    hint_texts = [_text(h) for h in hints]
    blob = " ".join([header, *hint_texts])

    is_key = bool(_SPREADSHEET_KEY_RE.search(header))
    required = header.endswith("*") or header.startswith("*")
    for hint in hint_texts:
        low = hint.casefold()
        if "optional" in low or "not required" in low:
            continue
        if "required" in low:
            required = True

    col_type = ColumnType.UNKNOWN
    for candidate, pattern in _TYPE_PATTERNS:
        if pattern.search(blob):
            col_type = candidate
            break

    ref_id_type = ""
    if col_type is ColumnType.REFERENCE:
        m = _REF_ID_TYPE_RE.search(blob)
        if m:
            ref_id_type = m.group(1)

    type_inferred = False
    if col_type is ColumnType.UNKNOWN:
        inferred = _infer_type(sample)
        if inferred is not ColumnType.UNKNOWN:
            col_type, type_inferred = inferred, True

    return ColumnSpec(
        index=index,
        header=header.rstrip("*").strip(),
        group=group.strip(),
        col_type=col_type,
        required=required,
        is_key=is_key,
        ref_id_type=ref_id_type,
        type_inferred=type_inferred,
        hints=hint_texts,
    )


def _infer_type(sample: list[Any]) -> ColumnType:
    values = [v for v in sample if v is not None and _text(v) != ""][:_INFER_SAMPLE]
    if not values:
        return ColumnType.UNKNOWN

    def check(pred: Any) -> bool:
        return all(pred(v) for v in values)

    if check(lambda v: isinstance(v, (dt.date, dt.datetime))):
        return ColumnType.DATE
    if check(lambda v: isinstance(v, str) and bool(_ISO_DATE_RE.match(v.strip()))):
        return ColumnType.DATE
    if check(lambda v: isinstance(v, bool)):
        return ColumnType.BOOLEAN
    if check(lambda v: isinstance(v, str) and v.strip().casefold() in _BOOL_TOKENS):
        return ColumnType.BOOLEAN
    if check(
        lambda v: (isinstance(v, (int, float)) and not isinstance(v, bool))
        or (isinstance(v, str) and bool(_NUMERIC_TEXT_RE.match(v.strip())))
    ):
        return ColumnType.NUMERIC
    return ColumnType.TEXT


# ---------------------------------------------------------------------------
# Grid utilities


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _is_blank_row(row: list[Any]) -> bool:
    return all(v is None or _text(v) == "" for v in row)


def _used_width(row: list[Any]) -> int:
    width = 0
    for i, v in enumerate(row):
        if v is not None and _text(v) != "":
            width = i + 1
    return width


def _trim_trailing_blank_rows(rows: list[list[Any]]) -> list[list[Any]]:
    end = len(rows)
    while end and _is_blank_row(rows[end - 1]):
        end -= 1
    return rows[:end]


def _trim_grid(grid: list[list[Any]]) -> list[list[Any]]:
    """Drop trailing all-blank rows and trailing all-blank columns."""
    grid = _trim_trailing_blank_rows(grid)
    width = max((_used_width(r) for r in grid), default=0)
    return [row[:width] for row in grid]
